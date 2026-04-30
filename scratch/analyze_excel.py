import openpyxl

# Analyze ESCRITURA
wb = openpyxl.load_workbook(r'e:\ANTIGRAVITY PROJECTS\PRUEBA SKILLS\INFORMACION BASE\ESCRITURA.xlsx')
ws = wb.active
print("=== ESCRITURA File ===")
print("Rows:", ws.max_row, "Cols:", ws.max_column)
print()

# Show first 5 rows
for row_idx in range(1, min(8, ws.max_row + 1)):
    vals = []
    for col_idx in range(1, ws.max_column + 1):
        v = ws.cell(row_idx, col_idx).value
        vals.append(str(v) if v is not None else 'None')
    row_str = " | ".join(vals)
    print("Row " + str(row_idx) + ": " + row_str)

print()

# Unique values for writing response columns (E, F, G)
for col_idx in [5, 6, 7]:
    uniques = set()
    for row_idx in range(1, ws.max_row + 1):
        v = ws.cell(row_idx, col_idx).value
        if v is not None:
            uniques.add(str(v))
    col_letter = chr(64 + col_idx)
    print("Col " + col_letter + " unique values: " + str(uniques))

print()
print()

# Analyze LECTURA
wb2 = openpyxl.load_workbook(r'e:\ANTIGRAVITY PROJECTS\PRUEBA SKILLS\INFORMACION BASE\LECTURA.xlsx')
ws2 = wb2.active
print("=== LECTURA File ===")
print("Rows:", ws2.max_row, "Cols:", ws2.max_column)
print()

for row_idx in range(1, min(5, ws2.max_row + 1)):
    vals = []
    for col_idx in range(1, ws2.max_column + 1):
        v = ws2.cell(row_idx, col_idx).value
        vals.append(str(v) if v is not None else 'None')
    row_str = " | ".join(vals)
    print("Row " + str(row_idx) + ": " + row_str)

print()
# Unique values in response columns
for col_idx in range(5, ws2.max_column + 1):
    uniques = set()
    for row_idx in range(1, ws2.max_row + 1):
        v = ws2.cell(row_idx, col_idx).value
        if v is not None:
            uniques.add(str(v))
    col_letter = chr(64 + col_idx)
    print("Col " + col_letter + " unique values: " + str(uniques))
