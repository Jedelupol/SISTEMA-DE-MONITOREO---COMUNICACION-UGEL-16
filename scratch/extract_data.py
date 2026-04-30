import openpyxl
import json

wb = openpyxl.load_workbook(r'e:/ANTIGRAVITY PROJECTS/PRUEBA SKILLS/INFORMACION BASE/SISTEMA_EVALUACION_DIAGNOSTICA_LIMA_PROVINCIAS.xlsx', read_only=True)

# Extract Parameters
params_sheet = wb['PARAMETROS']
ugels = []
institutions = []
for row in params_sheet.iter_rows(min_row=2, values_only=True):
    if row[0]: ugels.append(str(row[0]))
    if row[1]: institutions.append(str(row[1]))

ugels = sorted(list(set(ugels)))
institutions = sorted(list(set(institutions)))

# Extract Matrix data
matrix_sheet = wb['MATRIZ_CLAVES']
matrix_data = {}
# Columns: Grado (maybe?), Pregunta, Competencia, Capacidad, Desempeo, Clave, etc.
# Let's check headers first
headers = [cell.value for cell in matrix_sheet[1]]
print("Headers:", headers)

for row in matrix_sheet.iter_rows(min_row=2, values_only=True):
    # Depending on headers, we extract
    pass

data = {
    "ugels": ugels,
    "institutions": institutions
}

with open('e:/ANTIGRAVITY PROJECTS/PRUEBA SKILLS/edu-metrics-pro/lib/excel_data.json', 'w', encoding='utf-8') as f:
    json.dump(data, f, indent=2, ensure_ascii=False)

print("Data exported to excel_data.json")
